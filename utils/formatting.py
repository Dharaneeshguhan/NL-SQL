"""Currency, display, and export formatting helpers."""

from __future__ import annotations

import re
import textwrap
from io import BytesIO

import pandas as pd
from openpyxl.utils import get_column_letter

try:
    from fpdf import FPDF
except ImportError:  # pragma: no cover - optional export dependency
    FPDF = None

from config.constants import CURRENCY_LABEL

# Column name hints for currency formatting in result tables
_CURRENCY_COLUMNS = re.compile(
    r"(amount|sales|revenue|salary|price|total|cost|value|income)",
    re.IGNORECASE,
)


def format_inr(value: float | int | None, decimals: int = 2) -> str:
    """Format a number as Indian Rupees (Rs)."""
    if value is None:
        return f"{CURRENCY_LABEL} 0"
    return f"{CURRENCY_LABEL} {float(value):,.{decimals}f}"


def format_number(value: float | int | None, decimals: int = 0) -> str:
    """Format a plain number with thousands separators."""
    if value is None:
        return "0"
    if decimals == 0:
        return f"{int(value):,}"
    return f"{float(value):,.{decimals}f}"


def format_dataframe_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """
    Return a copy of df with currency columns formatted as Rs strings for UI display.
    """
    if df is None or df.empty:
        return df

    display = df.copy()
    for col in display.columns:
        if _CURRENCY_COLUMNS.search(col) and pd.api.types.is_numeric_dtype(display[col]):
            display[col] = display[col].apply(
                lambda v: format_inr(v) if pd.notna(v) else ""
            )
    return display


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Export DataFrame to CSV bytes for Streamlit download button."""
    return df.to_csv(index=False).encode("utf-8")


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Export DataFrame to an auto-sized Excel workbook."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        worksheet = writer.sheets["Results"]
        for idx, col in enumerate(df.columns, start=1):
            values = df[col].astype(str) if not df.empty else pd.Series(dtype=str)
            max_len = max([len(str(col)), *(values.map(len).tolist() or [0])]) + 2
            worksheet.column_dimensions[get_column_letter(idx)].width = min(max_len, 42)
    buffer.seek(0)
    return buffer.read()


def query_report_pdf_bytes(question: str, answer: str, sql: str, df: pd.DataFrame) -> bytes:
    """Create a compact PDF report for a completed query."""
    if FPDF is None:
        raise RuntimeError("PDF export is unavailable because fpdf is not installed.")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "NL-to-SQL Query Report")
    pdf.ln(12)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Question")
    pdf.ln(8)
    pdf.set_font("Helvetica", size=10)
    pdf.multi_cell(0, 6, _pdf_safe(question))
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Answer")
    pdf.ln(8)
    pdf.set_font("Helvetica", size=10)
    pdf.multi_cell(0, 6, _pdf_safe(answer))
    pdf.ln(2)

    if sql:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "SQL")
        pdf.ln(8)
        pdf.set_font("Helvetica", size=9)
        pdf.multi_cell(0, 5, _pdf_safe(textwrap.fill(sql.replace("\n", " "), width=95)))
        pdf.ln(2)

    if df is not None and not df.empty:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "Result Preview")
        pdf.ln(8)
        pdf.set_font("Helvetica", size=8)
        for _, row in df.head(20).iterrows():
            row_text = " | ".join(str(value) for value in row.tolist())
            pdf.multi_cell(0, 5, _pdf_safe(textwrap.fill(row_text, width=105)))

    output = pdf.output(dest="S")
    if isinstance(output, str):
        return output.encode("latin-1", errors="replace")
    return bytes(output)


def _pdf_safe(value: object) -> str:
    """FPDF core fonts are latin-1; replace unsupported characters gracefully."""
    return str(value).encode("latin-1", errors="replace").decode("latin-1")
