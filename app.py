"""NL-to-SQL Analytics Agent — Streamlit UI."""

import io
import os
import sys
import tempfile

# UTF-8 on Windows (prevents ascii codec errors with API / unicode text)
if sys.platform == "win32":
    os.environ.setdefault("PYTHONUTF8", "1")

import pandas as pd
import plotly.express as px
import streamlit as st
import textwrap
from openpyxl.utils import get_column_letter

try:
    import speech_recognition as sr
except ImportError:
    sr = None

try:
    import whisper
except ImportError:
    whisper = None

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

from agent import resolve_model, run_analytics_agent, explain_sql
from database import init_db, kpi_metrics, run_query, get_table_info

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NL-to-SQL Analytics Agent",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');
    html, body, [class*="css"] { font-family: 'Outfit', sans-serif; color: #1e1b4b; }
    .stApp {
        background: linear-gradient(165deg, #fff7ed 0%, #fdf4ff 40%, #f0fdfa 100%);
    }
    [data-testid="stSidebar"] {
        background: #ffffff;
        border-right: 1px solid #e9d5ff;
        box-shadow: 4px 0 24px rgba(124, 58, 237, 0.06);
    }
    [data-testid="stSidebar"] .stMarkdown h2 {
        color: #7c3aed !important;
        font-weight: 700;
    }
    [data-testid="stSidebar"] .stCaption { color: #6b7280; }

    .hero {
        background: linear-gradient(120deg, #7c3aed 0%, #a855f7 45%, #14b8a6 100%);
        border-radius: 20px;
        padding: 32px 36px;
        margin-bottom: 28px;
        box-shadow: 0 12px 40px rgba(124, 58, 237, 0.25);
        position: relative;
        overflow: hidden;
    }
    .hero::after {
        content: '';
        position: absolute; right: -20px; top: -20px;
        width: 200px; height: 200px;
        background: rgba(255,255,255,0.12);
        border-radius: 50%;
    }
    .hero h1 {
        margin: 0;
        font-size: 2rem;
        font-weight: 700;
        color: #ffffff;
        letter-spacing: -0.02em;
    }
    .hero p { color: rgba(255,255,255,0.92); margin: 10px 0 0; font-size: 1.05rem; max-width: 640px; }

    div[data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid #e9d5ff;
        border-radius: 16px;
        padding: 18px 22px;
        box-shadow: 0 4px 20px rgba(124, 58, 237, 0.08);
    }
    div[data-testid="stMetric"] label { color: #6b7280 !important; font-weight: 500; }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #7c3aed !important;
        font-weight: 700;
    }

    .bubble-tag {
        font-size: 0.7rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        opacity: 0.85;
        margin-right: 6px;
    }
    .user-bubble {
        background: linear-gradient(135deg, #7c3aed, #a855f7);
        color: #ffffff;
        padding: 14px 20px;
        border-radius: 20px 20px 4px 20px;
        margin: 10px 0 10px 40px;
        font-size: 0.95rem;
        line-height: 1.55;
        box-shadow: 0 6px 20px rgba(124, 58, 237, 0.25);
    }
    .assistant-bubble {
        background: #ffffff;
        color: #374151;
        padding: 14px 20px;
        border-radius: 20px 20px 20px 4px;
        margin: 10px 40px 10px 0;
        font-size: 0.95rem;
        line-height: 1.55;
        border: 1px solid #e5e7eb;
        box-shadow: 0 2px 12px rgba(0,0,0,0.04);
    }
    .assistant-bubble .bubble-tag { color: #7c3aed; opacity: 1; }
    .sql-label {
        font-size: 0.68rem;
        text-transform: uppercase;
        letter-spacing: 0.12em;
        color: #9ca3af;
        margin: 14px 0 6px;
        font-weight: 600;
    }
    .sql-block {
        background: #1e1b4b;
        border-left: 4px solid #14b8a6;
        border-radius: 10px;
        padding: 14px 18px;
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.8rem;
        color: #5eead4;
        white-space: pre-wrap;
        word-break: break-word;
    }
    .tool-pill {
        display: inline-block;
        background: #f3e8ff;
        border: 1px solid #d8b4fe;
        border-radius: 999px;
        padding: 3px 12px;
        font-size: 0.72rem;
        color: #6b21a8;
        margin: 2px 6px 2px 0;
        font-weight: 500;
    }
    .stTextInput input, .stSelectbox > div > div {
        background: #ffffff !important;
        border: 1px solid #d8b4fe !important;
        border-radius: 12px !important;
        color: #1e1b4b !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #7c3aed, #14b8a6) !important;
        border: none !important;
        border-radius: 12px !important;
        font-weight: 600 !important;
        color: white !important;
    }
    .stButton > button[kind="primary"]:hover {
        box-shadow: 0 6px 20px rgba(124, 58, 237, 0.35) !important;
    }
    hr { border-color: #e9d5ff !important; }
    [data-testid="stDataFrame"] { border-radius: 12px; border: 1px solid #e5e7eb; }
</style>
""",
    unsafe_allow_html=True,
)


def make_chart(df: pd.DataFrame, chart_type: str, x_col: str, y_col: str, title: str):
    if not chart_type or x_col not in df.columns or y_col not in df.columns:
        return None
    palette = ["#7c3aed", "#a855f7", "#14b8a6", "#f97316", "#ec4899"]
    builders = {
        "bar": lambda: px.bar(df, x=x_col, y=y_col, title=title, color_discrete_sequence=palette),
        "line": lambda: px.line(
            df, x=x_col, y=y_col, title=title, markers=True, color_discrete_sequence=["#7c3aed"]
        ),
        "pie": lambda: px.pie(df, names=x_col, values=y_col, title=title, color_discrete_sequence=palette),
        "scatter": lambda: px.scatter(
            df, x=x_col, y=y_col, title=title, color_discrete_sequence=["#14b8a6"]
        ),
    }
    builder = builders.get(chart_type)
    if not builder:
        return None
    fig = builder()
    fig.update_layout(
        plot_bgcolor="#ffffff",
        paper_bgcolor="#faf5ff",
        font_color="#374151",
        title_font_color="#7c3aed",
        title_font_size=15,
        xaxis=dict(gridcolor="#f3e8ff", tickfont_color="#6b7280"),
        yaxis=dict(gridcolor="#f3e8ff", tickfont_color="#6b7280"),
        margin=dict(l=48, r=24, t=56, b=48),
        height=380,
    )
    return fig


def _df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        worksheet = writer.sheets["Results"]
        for idx, col in enumerate(df.columns, start=1):
            column_len = max(
                df[col].astype(str).map(len).max(),
                len(str(col)),
            ) + 2
            worksheet.column_dimensions[get_column_letter(idx)].width = min(column_len, 40)
    buffer.seek(0)
    return buffer.read()


def _load_whisper_model():
    if whisper is None:
        return None
    try:
        return whisper.load_model("tiny")
    except Exception:
        return None

_whisper_model = None


def speech_to_text() -> str:
    if sr is None:
        st.error("SpeechRecognition is not installed.")
        return ""

    recognizer = sr.Recognizer()

    try:
        with sr.Microphone() as source:
            st.info("🎤 Listening... Speak now")
            recognizer.adjust_for_ambient_noise(source, duration=1)
            audio = recognizer.listen(
                source,
                timeout=5,
                phrase_time_limit=10,
            )

        st.info("🔄 Converting speech to text...")

        if whisper is not None:
            global _whisper_model
            if _whisper_model is None:
                _whisper_model = _load_whisper_model()
            if _whisper_model:
                wav_bytes = audio.get_wav_data(convert_rate=16000, convert_width=2)
                with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as tmp:
                    tmp.write(wav_bytes)
                    tmp_path = tmp.name
                try:
                    result = _whisper_model.transcribe(tmp_path)
                    return result.get("text", "").strip()
                finally:
                    os.remove(tmp_path)

        text = recognizer.recognize_google(audio)
        return text

    except sr.WaitTimeoutError:
        st.error("No speech detected.")
        return ""

    except sr.UnknownValueError:
        st.error("Could not understand audio.")
        return ""

    except Exception as e:
        st.error(f"Voice Error: {e}")
        return ""


def _sql_report_pdf(question: str, answer: str, sql: str, df: pd.DataFrame) -> bytes:
    if FPDF is None:
        raise RuntimeError("PDF export is unavailable because the FPDF library is not installed.")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "NL-to-SQL Query Report")
    pdf.ln(10)
    pdf.set_font("Helvetica", size=11)
    pdf.multi_cell(0, 8, f"Question: {question}")
    pdf.ln(1)
    pdf.multi_cell(0, 8, f"Answer: {answer}")
    pdf.ln(1)
    pdf.multi_cell(0, 8, "SQL Query:")
    pdf.set_font("Helvetica", size=10)
    safe_sql = textwrap.fill(sql.replace("\n", " "), width=90, break_long_words=True, replace_whitespace=False)
    pdf.multi_cell(0, 6, safe_sql)
    if not df.empty:
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Result preview:")
        pdf.ln(8)
        pdf.set_font("Helvetica", size=9)
        preview = df.head(20)
        for _, row in preview.iterrows():
            row_text = " | ".join(str(value) for value in row.tolist())
            safe_row = textwrap.fill(row_text, width=90, break_long_words=True, replace_whitespace=False)
            if pdf.get_y() > pdf.h - 30:
                pdf.add_page()
            pdf.multi_cell(0, 6, safe_row)
    output = pdf.output(dest="S")
    if isinstance(output, str):
        return output.encode("latin-1")
    return output


def _create_export_buttons(idx: int, item: dict) -> None:
    if item.get("df") is None or item["df"].empty:
        return
    excel_bytes = _df_to_excel_bytes(item["df"])
    pdf_bytes = None
    try:
        pdf_bytes = _sql_report_pdf(
            item.get("question", ""),
            item.get("content", ""),
            item.get("sql", ""),
            item["df"],
        )
    except Exception:
        pdf_bytes = None

    st.download_button(
        label="Download Excel",
        data=excel_bytes,
        file_name=f"query_results_{idx + 1}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_xlsx_{idx}",
    )
    if pdf_bytes is not None:
        st.download_button(
            label="Download PDF",
            data=pdf_bytes,
            file_name=f"query_report_{idx + 1}.pdf",
            mime="application/pdf",
            key=f"download_pdf_{idx}",
        )
    else:
        st.markdown("*PDF export unavailable for this query.*")


if "history" not in st.session_state:
    st.session_state.history = []
if "query_history" not in st.session_state:
    st.session_state.query_history = []
if "db_ready" not in st.session_state:
    init_db()
    st.session_state.db_ready = True

OPENROUTER_MODELS = [
    "google/gemini-2.5-flash",
    "google/gemini-2.5-flash-lite",
    "google/gemini-3.5-flash",
    "meta-llama/llama-3.3-70b-instruct:free",
    "google/gemma-4-26b-a4b-it:free",
    "moonshotai/kimi-k2.6:free",
    "qwen/qwen-2.5-7b-instruct",
]

SUGGESTIONS = [
    "Top 5 customers by revenue",
    "Monthly sales trend in 2024",
    "Revenue by product category",
    "Products with low stock (under 20)",
    "Average order value by customer segment",
    "Count of orders by status",
]

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🗄️ Analytics Agent")
    st.caption("Schema-grounded NL → SQL")

    api_key = st.text_input(
        "OpenRouter API Key",
        type="password",
        value=os.environ.get("OPENROUTER_API_KEY", ""),
        placeholder="sk-or-... or set OPENROUTER_API_KEY",
        help="Get a key at https://openrouter.ai/keys",
    )
    default_model = resolve_model(os.environ.get("OPENROUTER_MODEL"))
    model_idx = (
        OPENROUTER_MODELS.index(default_model)
        if default_model in OPENROUTER_MODELS
        else 0
    )
    model = st.selectbox("Model", OPENROUTER_MODELS, index=model_idx)
    st.caption("Powered by **[OpenRouter](https://openrouter.ai)** — pick a free or paid model")
    st.markdown("---")

    st.markdown("### 📋 Schema explorer")
    for tname, tdata in get_table_info().items():
        with st.expander(f"**{tname}** · {tdata['count']} rows"):
            for col in tdata["columns"]:
                pk = " 🔑" if col[5] else ""
                st.markdown(f"`{col[1]}` · *{col[2]}*{pk}")

    st.markdown("---")
    st.markdown("### 💡 Example questions")
    for s in SUGGESTIONS:
        if st.button(s, key=f"sug_{s}", width="stretch"):
            st.session_state["prefill"] = s

    st.markdown("---")
    st.markdown("### 🎤 Voice-to-SQL")
    st.caption(
        "Speak naturally:\n\n"
        "- Top customers by revenue\n"
        "- Revenue by category\n"
        "- Monthly sales trend"
    )
    if whisper is not None:
        st.caption("Local Whisper transcription is available if installed.")

    st.markdown("---")
    st.markdown("### 🕘 Query History")
    if st.session_state.query_history:
        for idx, item in enumerate(reversed(st.session_state.query_history[-10:])):
            history_key = f"history_use_{len(st.session_state.query_history)-idx-1}"
            if st.button(item["question"], key=history_key, width="stretch"):
                st.session_state["prefill"] = item["question"]
    else:
        st.info("No previous queries yet.")

    st.markdown("---")
    if st.button("🗑️ Clear conversation", width="stretch"):
        st.session_state.history = []
        st.session_state.query_history = []
        st.rerun()

# ── Main ──────────────────────────────────────────────────────────────────────
st.markdown(
    """
<div class="hero">
  <h1>NL-to-SQL Analytics Agent</h1>
  <p>Ask in plain English. The agent inspects your SQLite schema with tools,
  writes safe read-only SQL, and returns answers with charts when it helps.</p>
</div>
""",
    unsafe_allow_html=True,
)

kpis = kpi_metrics()
c1, c2, c3, c4 = st.columns(4)
c1.metric("Completed orders", kpis["orders"])
c2.metric("Total revenue", f"${kpis['revenue']:,.2f}")
c3.metric("Customers", kpis["customers"])
c4.metric("Products", kpis["products"])

st.markdown("---")

for idx, item in enumerate(st.session_state.history):
    if item["role"] == "user":
        st.markdown(
            f'<div class="user-bubble"><span class="bubble-tag">You</span> {item["content"]}</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div class="assistant-bubble"><span class="bubble-tag">Agent</span> {item["content"]}</div>',
            unsafe_allow_html=True,
        )
        if item.get("tools"):
            pills = "".join(
                f'<span class="tool-pill">{t["tool"]}</span>' for t in item["tools"]
            )
            st.markdown(f"**Agent tools used:** {pills}", unsafe_allow_html=True)
        if item.get("sql"):
            st.markdown(
                f'<div class="sql-label">Generated SQL</div>'
                f'<div class="sql-block">{item["sql"]}</div>',
                unsafe_allow_html=True,
            )
            if st.button("Explain SQL", key=f"explain_sql_{idx}"):
                try:
                    explanation = explain_sql(item["sql"], api_key, model=model)
                    st.session_state.history[idx]["sql_explanation"] = explanation
                except Exception as exc:
                    st.session_state.history[idx]["sql_explanation"] = f"Error: {exc}"
            if item.get("sql_explanation"):
                st.markdown(
                    f'**SQL explanation:**<br>{item["sql_explanation"]}',
                    unsafe_allow_html=True,
                )
        if item.get("error"):
            st.error(item["error"])
        elif item.get("df") is not None and not item["df"].empty:
            st.dataframe(item["df"], width="stretch", hide_index=True)
            _create_export_buttons(idx, item)
        if item.get("chart") is not None:
            st.plotly_chart(item["chart"], width="stretch")

st.markdown("---")
prefill = st.session_state.pop("prefill", "")
col_q, col_voice, col_run = st.columns([4, 1, 1])
with col_q:
    question = st.text_input(
        "Question",
        value=prefill,
        placeholder="e.g. Which product categories drive the most revenue?",
        label_visibility="collapsed",
    )
with col_voice:
    voice_input = st.button("🎤 Voice", width="stretch")

with col_run:
    submit = st.button("▶ Ask", type="primary", width="stretch")

if voice_input:
    spoken_text = speech_to_text()
    if spoken_text:
        st.session_state["prefill"] = spoken_text
        st.session_state["spoken_text"] = spoken_text
        st.rerun()

if st.session_state.get("spoken_text"):
    st.success(f"Recognized: {st.session_state['spoken_text']}")

if submit and question.strip():
    if not api_key:
        st.warning("Add your OpenRouter API key in the sidebar or set `OPENROUTER_API_KEY`.")
    else:
        st.session_state.history.append({"role": "user", "content": question})
        with st.spinner("Agent is exploring schema and building SQL…"):
            try:
                out = run_analytics_agent(question, api_key, model=model)
                result = out["result"]
                sql = result.get("sql", "")
                df, err = run_query(sql) if sql else (pd.DataFrame(), None)
                chart = None
                if df is not None and not df.empty and result.get("chart_type"):
                    chart = make_chart(
                        df,
                        result["chart_type"],
                        result.get("chart_x", ""),
                        result.get("chart_y", ""),
                        result.get("chart_title", "Chart"),
                    )
                assistant_item = {
                    "role": "assistant",
                    "question": question,
                    "content": result.get("answer", "Here are your results."),
                    "sql": sql,
                    "df": df if df is not None else pd.DataFrame(),
                    "chart": chart,
                    "error": err,
                    "tools": out.get("tool_trace", []),
                }
                st.session_state.history.append(assistant_item)
                st.session_state.query_history.append(
                    {
                        "question": question,
                        "sql": sql,
                        "answer": assistant_item["content"],
                    }
                )
            except Exception as exc:
                st.session_state.history.append(
                    {
                        "role": "assistant",
                        "content": "I could not complete that request.",
                        "sql": None,
                        "df": None,
                        "chart": None,
                        "error": str(exc),
                        "tools": [],
                    }
                )
        st.rerun()
